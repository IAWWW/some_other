import os
import requests
import openpyxl
import re
import openai
import time

# Use your API key here
openai.api_key = "..."

# Load the workbook
wb = openpyxl.load_workbook('/Users/iar/Documents/excel/data_base.xlsx')

save = '/Users/iar/Documents/excel/data_base19.xlsx'

# Select the first sheet in the workbook
worksheet = wb.worksheets[3]


counter = 0
for row in worksheet.iter_rows():
    if 1 < counter <= 3:
        print(counter)
    
        row_index=row[0].row
        brand = str(worksheet["A"+str(row_index)].value)
        model = str(worksheet["C"+ str(row_index)].value)
        #trim = str(worksheet["C"+ str(row_index)].value)

        # Define the prompt you want GPT-3 to complete
        prompt = "give me caracteristique of this yacht : "+brand+" "+model+", I want Max speed in knots, Engine, Length in meter, Range in NM, Year, Price in USD"

        # Sleep
        print("sleep 10")
        time.sleep(10)

        # Use the completions API to generate text
        completions = openai.Completion.create(
            engine="text-davinci-003",
            prompt=prompt,
            max_tokens=200,
            n=1,
            stop=None,
            temperature=0.1,
        )
        pattern_speed = r'Max.*:?(\d[0-9])'
        pattern_engine = r'Engin.*: ?(.*)'
        pattern_length = r'Leng.*: ?(\d+.?\d+)'
        pattern_range = r'Rang.*: ?(\d,?\d+)'
        pattern_price = r'Price ?: ?(.*) '
        pattern_year = r'Year.*: ?(\d+)'

        # Print the generated text
        message = completions.choices[0].text

        match_speed = re.search(pattern_speed, message)
        match_engine = re.search(pattern_engine, message)
        match_lenght = re.search(pattern_length, message)
        match_range = re.search(pattern_range, message)
        match_price = re.search(pattern_price, message)
        match_year = re.search(pattern_year, message)

        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=4)
        if match_speed:
            
            matching_string = match_speed.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for mouvement", message)
            cell.value = "-"
            wb.save(save)
        
        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=5)
        if match_engine:
            matching_string = match_engine.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for case", message)
            cell.value = "-"
            wb.save(save)
        
        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=6)
        if match_lenght:
            
            matching_string = match_lenght.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for size", message)
            cell.value = "-"
            wb.save(save)
        
        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=7)
        if match_range:
            
            matching_string = match_range.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for year", message)
            cell.value = "-"
            wb.save(save)
        
        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=9)
        if match_price:
            matching_string = match_price.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for price", message)
            cell.value = "-"
            wb.save(save)

        # Cellule à editer
        cell = worksheet.cell(row=row_index, column=8)
        if match_year:
            matching_string = match_year.group(1)
            print(matching_string)
            cell.value = matching_string
            wb.save(save)
        else:
            print("No match for year", message)
            cell.value = "-"
            wb.save(save)
    counter += 1
# Close the workbook
wb.close()



