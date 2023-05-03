import requests
import openpyxl
import os
import re
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup

hdr = {'User-Agent': 'Mozilla/5.0'}

# Load the workbook
wb = openpyxl.load_workbook('/Users/iar/Documents/.../data_base_watch.xlsx')
excel_save = '/Users/iar/Documents/.../data_base_99.xlsx'
# Select the active worksheet
ws = wb.active

url = 'https://askme.watch'
# Select the sheet in the workbook
worksheet = wb.worksheets[2]




counter = 0
for row in worksheet.iter_rows():
    x = 0

    if 0 < counter <= 133:
        print(counter)

        row_index=row[0].row
        brand = str(worksheet["A"+str(row_index)].value).lower().replace(" ","-")
        collection = str(worksheet["B"+ str(row_index)].value).lower().replace(" ","-")
        model = str(worksheet["C"+ str(row_index)].value).lower().replace("/","").replace(" ","-")
        
        brand_search = str(worksheet["A"+str(row_index)].value).replace(" ","%20")
        collection_search = str(worksheet["B"+ str(row_index)].value).replace(" ","%20")
        model_search = str(worksheet["C"+ str(row_index)].value).replace(" ","%20")
        site = "https://askme.watch/en/watch-finder-search?search_fulltext="+collection_search+"%20"+model_search
        r = Request(site,headers=hdr)
        try:
            page = urlopen(r)
            soup = BeautifulSoup(page, 'html.parser')
            for item in soup.findAll('a'):
                url_href = item['href']
                
                if "%" not in str(url_href):
                    if model in str(url_href) and brand in str(url_href) and collection in str(url_href):

                        url_watche = url+url_href
                        r2 = Request(url_watche,headers=hdr)
                        
                        try:
                            page2 = urlopen(r2)
                            soup2 = BeautifulSoup(page2, 'html.parser')

                            # put brand
                            cell_brand = worksheet.cell(row=row_index+143, column=1)
                            cell_brand.value = str(worksheet["A"+str(row_index)].value)
                            wb.save(excel_save)

                            # find collection + model
                            coloc_model = soup2.find('div', {'class':'wdr-header'})

                            text_collection = coloc_model.find_next("a").text
                            text_model = coloc_model.find_next("h1").text
                            
                            if text_collection and text_model:
                                # Cellule à editer
                                cell = worksheet.cell(row=row_index+143, column=2)
                                cell.value = text_collection
                                
                                cell2 = worksheet.cell(row=row_index+143, column=3)
                                cell2.value = text_model
                                wb.save(excel_save)

                            # Dans le nouveau lien chercher photo url
                            for img in soup2.findAll('div', {'class':'watch-detail_left-big-image'}):
                                brandee = str(worksheet["A"+str(row_index)].value)
                                print(brandee, text_collection, text_model)
                                for image in img.findAll('img'):
                                    x += 1
                                    url_photo = image.get("src")
                                    print(brandee, text_collection, text_model, url_photo)
                                    try:
                                        with open(brandee.replace(" ","-").lower()+"_"+text_collection.lower().replace(" ","-")+"_"+text_model.replace("/","").replace(" ","").lower()+"_"+str(x)+".jpg", "wb") as f:
                                            f.write(requests.get(url_photo).content)
                                        # Cellule à editer
                                        cell = worksheet.cell(row=row_index+143, column=12)
                                        cell.value = "found"
                                        print("ok")
                                        wb.save(excel_save)
                                    except:
                                        # Cellule à editer
                                        print("a")
                                        cell = worksheet.cell(row=row_index+143, column=12)
                                        cell.value = "notfound"
                                        wb.save(excel_save)
                            
                            

                            # find price
                            pricee = soup2.find('div', {'class':'wdr-bottom_item'})

                            text_price = pricee.find_next("h3").text
                            text_price = text_price.replace("€", "")
                            if text_price:
                                # Cellule à editer
                                cell_price = worksheet.cell(row=row_index+143, column=8)
                                cell_price.value = text_price
                                wb.save(excel_save)


                            # get info case material and put in excel
                            span_case = soup2.find("span", string="CASE MATERIAL")
                            text_case = span_case.find_next("strong").text
                            if text_case:
                                # Cellule à editer
                                cell_case = worksheet.cell(row=row_index+143, column=5)
                                cell_case.value = text_case
                                wb.save(excel_save)

                            # get info movement and put in excel
                            span_movement = soup2.find("span", string="Movement type")
                            text_movement = span_movement.find_next("strong").text
                            if text_movement:
                                # Cellule à editer
                                cell_mov = worksheet.cell(row=row_index+143, column=4)
                                cell_mov.value = text_movement
                                wb.save(excel_save)

                            # get info size and put in excel
                            span_size = soup2.find("span", string="CASE DIAMETER")
                            text_size = span_size.find_next("strong").text
                            if text_size:
                                # Cellule à editer
                                cell_size = worksheet.cell(row=row_index+143, column=6)
                                cell_size.value = text_size
                                wb.save(excel_save)


                                
                        except:
                            print("fail get url", url_watche, brand, collection, model)
        except:
            print("fail url for ", brand, collection, model)
            

    counter += 1
# Close the workbook
wb.close()
