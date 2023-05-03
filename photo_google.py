import os
import openpyxl
import time
from google_images_search import GoogleImagesSearch

gis = GoogleImagesSearch('...', '...')
# https://github.com/arrrlo/Google-Images-Search

# Load the workbook
wb = openpyxl.load_workbook('/Users/iar/Documents/excel/photo_manq.xlsx')

# Select the active worksheet
ws = wb.worksheets[1]

for i in range (120,179): #a faire
    time.sleep(1)
    search_params = { ... }
    brand = str(ws["A"+str(i)].value)
    model = str(ws["B"+ str(i)].value)
    trim = str(ws["C"+ str(i)].value)

    brand_name = brand.replace(' ','-').replace('/',"").lower()
    model_name = model.replace(' ','-').replace('/',"").lower()
    trim_name = trim.replace(' ','-').replace('/',"").lower()
    print(brand + ' '+ model + ' ' + trim)
    # Search google + dl 
    search_params = {
        'q': brand + ' '+ model + ' ' + trim,
        'num': 3,
        'fileType': 'jpg', #png
        'imgType': 'photo', ##
        'imgSize': 'xlarge',
        #'imgSize': 'huge|icon|large|medium|small|xlarge|xxlarge|imgSizeUndefined',
    }

    gis.search(search_params=search_params, 
        path_to_dir='/Users/iar/Documents/.../python/images_cars_google', 
        custom_image_name= brand_name+'_'+model_name+'_'+trim_name+'_',
    )

    # Cellule à editer
    cell = ws.cell(row=i, column=4)
    cell.value = 'oui'
    # Save the changes to the workbook
    wb.save('/Users/iar/Documents/.../photo_manq.xlsx')
# Close the workbook
wb.close()

